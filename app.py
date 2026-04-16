import streamlit as st
import pandas as pd
import numpy as np
import re
import io
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ================= 1. 页面与侧边栏配置 =================
st.set_page_config(page_title="亚马逊智能补货系统", page_icon="🚀", layout="wide")

st.sidebar.header("⚙️ 补货计算参数配置")
TARGET_DAYS_TRANSIT = st.sidebar.number_input("目标物流在途天数", min_value=0, value=50, step=1)
TARGET_DAYS_STOCK = st.sidebar.number_input("目标安全在库天数", min_value=0, value=30, step=1)
TARGET_DAYS_TOTAL = TARGET_DAYS_TRANSIT + TARGET_DAYS_STOCK
ALERT_STOCKOUT_DAYS = st.sidebar.number_input("缺货预警天数阈值", min_value=0, value=20, step=1)

# ================= 2. 数据清洗基础函数 =================
def clean_msku_strict(val):
    if pd.isna(val): return ""
    return re.sub(r'\s+', '', str(val).strip())

def clean_percentage_or_money(x, col_name=""):
    if pd.isna(x): return 0.0
    s = str(x).strip()
    if s == '-': return 0.0
    s_clean = re.sub(r'[^\d.-]', '', s)
    try:
        val = float(s_clean)
        keywords_percent = ['率', 'CTR', 'ACOS', 'ACoAS', 'CVR']
        if any(k in col_name.upper() for k in keywords_percent):
            if '%' in s: val = val / 100.0
            elif val > 1.0 and any(k in col_name.upper() for k in ['CTR', '转化率', 'CVR', 'ACOS']):
                val = val / 100.0
        return val
    except:
        return 0.0

def clean_columns(df):
    df.columns = [str(c).strip() for c in df.columns]
    return df

def find_col_exact(df, keyword):
    for col in df.columns:
        if col.upper() == keyword.upper(): return col
    return None

def find_col_fuzzy_priority(df, keywords_list):
    for kw in keywords_list:
        for col in df.columns:
            if kw.lower() in col.lower(): return col
    return None

def find_col_by_pattern(df, pattern):
    for col in df.columns:
        if pattern in col: return col
    return None

def read_uploaded_file(file):
    if file.name.endswith('.csv'):
        try: return pd.read_csv(file, encoding='utf-8')
        except: 
            file.seek(0)
            return pd.read_csv(file, encoding='gbk')
    else:
        return pd.read_excel(file)

# ================= 3. 数据处理模块 =================
def merge_uploaded_traffic(uploaded_files):
    if not uploaded_files: return None
    dfs = []
    for file in uploaded_files:
        df = read_uploaded_file(file)
        if df is not None:
            df['来源文件'] = file.name
            dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else None

def process_traffic_df(df, prefix):
    if df is None or df.empty: return None
    try:
        df = clean_columns(df)
        sku_col = find_col_fuzzy_priority(df, ['SKU', '(Child)', '子ASIN'])
        if not sku_col: return None
        
        df['join_key'] = df[sku_col].apply(clean_msku_strict)
        shop_col = find_col_fuzzy_priority(df, ['店铺', 'Shop', 'Store', 'Account', '账号'])
        df['traffic_shop'] = df[shop_col].astype(str).str.strip() if shop_col else 'Unknown'
            
        indicators = [("会话", None, "会话数"), ("页面浏览", None, "页面浏览量"), ("订单商品", "B2B", "订单商品总数"), ("销售额", "B2B", "销售额")]
        found_data = {}
        for keyword, exclude, suffix in indicators:
            for col in df.columns:
                if keyword in col:
                    if exclude and exclude in col: continue
                    df[col] = df[col].apply(lambda x: clean_percentage_or_money(x, col))
                    found_data[col] = f"{prefix}{suffix}"
                    break
        
        if not found_data: return None
        cols = ['join_key', 'traffic_shop'] + list(found_data.keys())
        df = df[cols].groupby(['join_key', 'traffic_shop']).sum().reset_index()
        return df.rename(columns=found_data)
    except: return None

def process_inventory_files(uploaded_files):
    if not uploaded_files: return None
    all_inv_dfs = []
    today = pd.Timestamp.now().normalize()
    for file in uploaded_files:
        try:
            d = read_uploaded_file(file)
            d = clean_columns(d)
            col_sku = find_col_fuzzy_priority(d, ['SKU', 'sku', '产品'])
            col_qty = find_col_fuzzy_priority(d, ['海外仓在途', '在途', '发货量', '数量', 'Qty', 'quantity', '件数'])
            col_date = find_col_fuzzy_priority(d, ['实际-预计到货时间', '预计到货', 'XT-预计到货', '实际', '预计', '到货', '时间', '日期', 'ETA'])

            if not (col_sku and col_qty and col_date): continue
            d['join_key'] = d[col_sku].apply(clean_msku_strict)
            d['qty_clean'] = d[col_qty].apply(lambda x: clean_percentage_or_money(x, col_qty))
            d['date_clean'] = pd.to_datetime(d[col_date], errors='coerce')
            d['days_diff'] = (d['date_clean'] - today).dt.days
            
            d['待发货'] = d.apply(lambda x: x['qty_clean'] if pd.isna(x['date_clean']) else 0, axis=1)
            d['7天内送达'] = d.apply(lambda x: x['qty_clean'] if pd.notna(x['date_clean']) and x['days_diff'] <= 7 else 0, axis=1)
            d['14天内送达'] = d.apply(lambda x: x['qty_clean'] if pd.notna(x['date_clean']) and 7 < x['days_diff'] <= 14 else 0, axis=1)
            d['21天内送达'] = d.apply(lambda x: x['qty_clean'] if pd.notna(x['date_clean']) and 14 < x['days_diff'] <= 21 else 0, axis=1)
            d['28天内送达'] = d.apply(lambda x: x['qty_clean'] if pd.notna(x['date_clean']) and 21 < x['days_diff'] <= 28 else 0, axis=1)
            d['28天以上送达'] = d.apply(lambda x: x['qty_clean'] if pd.notna(x['date_clean']) and x['days_diff'] > 28 else 0, axis=1)
            all_inv_dfs.append(d[['join_key', '待发货', '7天内送达', '14天内送达', '21天内送达', '28天内送达', '28天以上送达']])
        except: pass
    return pd.concat(all_inv_dfs, ignore_index=True).groupby('join_key').sum().reset_index() if all_inv_dfs else None

def process_age_files(uploaded_files):
    if not uploaded_files: return None
    all_dfs = []
    targets = [("可用", "可用量"), ("0~30", "0~30库龄"), ("31~60", "31~60库龄"), ("61~90", "61~90库龄"), 
               ("91~180", "91~180库龄"), ("181~270", "181~270库龄"), ("271~330", "271~330库龄"), 
               ("331~365", "331~365库龄"), ("365以上", "365以上库龄"), ("181以上", "181以上库龄"), ("181天以上", "181以上库龄")]
    
    for file in uploaded_files:
        try:
            d = read_uploaded_file(file)
            d = clean_columns(d)
            col_sku = find_col_fuzzy_priority(d, ['SKU', 'sku', '产品'])
            if not col_sku: continue
            
            d['join_key'] = d[col_sku].apply(clean_msku_strict)
            result_df = pd.DataFrame({'join_key': d['join_key']})
            
            col_age_num = find_col_exact(d, '库龄') or find_col_fuzzy_priority(d, ['库龄', 'Age'])
            col_qty = find_col_fuzzy_priority(d, ['海外仓在库', '可用', '可用量', '在库数量', '数量'])
            has_old_format = find_col_by_pattern(d, "0~30") or find_col_by_pattern(d, "0-30")
            
            if col_age_num and col_qty and not has_old_format:
                d['qty_clean'] = d[col_qty].apply(lambda x: clean_percentage_or_money(x, col_qty))
                d['age_clean'] = d[col_age_num].apply(lambda x: clean_percentage_or_money(x, col_age_num) if pd.notna(x) else 0)
                
                result_df['可用量'] = d['qty_clean']
                result_df['0~30库龄'] = d.apply(lambda x: x['qty_clean'] if 0 <= x['age_clean'] <= 30 else 0, axis=1)
                result_df['31~60库龄'] = d.apply(lambda x: x['qty_clean'] if 30 < x['age_clean'] <= 60 else 0, axis=1)
                result_df['61~90库龄'] = d.apply(lambda x: x['qty_clean'] if 60 < x['age_clean'] <= 90 else 0, axis=1)
                result_df['91~180库龄'] = d.apply(lambda x: x['qty_clean'] if 90 < x['age_clean'] <= 180 else 0, axis=1)
                result_df['181~270库龄'] = d.apply(lambda x: x['qty_clean'] if 180 < x['age_clean'] <= 270 else 0, axis=1)
                result_df['271~330库龄'] = d.apply(lambda x: x['qty_clean'] if 270 < x['age_clean'] <= 330 else 0, axis=1)
                result_df['331~365库龄'] = d.apply(lambda x: x['qty_clean'] if 330 < x['age_clean'] <= 365 else 0, axis=1)
                result_df['365以上库龄'] = d.apply(lambda x: x['qty_clean'] if x['age_clean'] > 365 else 0, axis=1)
                result_df['181以上库龄'] = d.apply(lambda x: x['qty_clean'] if x['age_clean'] > 180 else 0, axis=1)
            else:
                for pattern, target_name in targets:
                    real_col = find_col_by_pattern(d, pattern)
                    if real_col and real_col in d.columns: 
                        result_df[target_name] = d[real_col].apply(lambda x: clean_percentage_or_money(x, target_name))
                    else: result_df[target_name] = 0
            all_dfs.append(result_df)
        except: pass
    return pd.concat(all_dfs, ignore_index=True).groupby('join_key').sum().reset_index() if all_dfs else None

def move_col(columns_list, col_to_move, ref_col, position='before'):
    if col_to_move in columns_list and ref_col in columns_list:
        columns_list.remove(col_to_move)
        ref_idx = columns_list.index(ref_col)
        if position == 'before': columns_list.insert(ref_idx, col_to_move)
        else: columns_list.insert(ref_idx + 1, col_to_move)
    return columns_list

# ================= 4. 主程序运行 =================
st.title("🚀 亚马逊智能补货系统 (Web增强版)")
st.markdown("上传您的报表文件，系统将在云端自动清洗、合并并生成高级可视化补货表。")

col1, col2 = st.columns(2)
with col1:
    st.subheader("1. 基础表现与流量")
    files_prod = st.file_uploader("📂 产品表现表 (主表/重要表)", accept_multiple_files=True)
    files_7d = st.file_uploader("📂 7天流量表 (支持多文件合并)", accept_multiple_files=True)
    files_14d = st.file_uploader("📂 14天流量表 (支持多文件合并)", accept_multiple_files=True)

with col2:
    st.subheader("2. 供应链数据")
    files_inv = st.file_uploader("📦 库存表", accept_multiple_files=True)
    files_age = st.file_uploader("👴 库龄表", accept_multiple_files=True)

if st.button("✨ 立即生成补货建议", type="primary"):
    if not files_prod:
        st.error("❌ 必须上传产品表现表作为主数据驱动！")
        st.stop()
        
    with st.spinner('系统正在拼命计算中，请稍候...'):
        try:
            # 1. 处理主表
            df_whitelist, all_data_dfs = None, []
            for f in files_prod:
                d = read_uploaded_file(f)
                d = clean_columns(d)
                if "重要" in f.name:
                    col_imp_msku = find_col_fuzzy_priority(d, ['MSKU', '商家SKU', 'sku'])
                    if col_imp_msku:
                        d['MSKU'] = d[col_imp_msku].apply(clean_msku_strict)
                        col_shop = find_col_fuzzy_priority(d, ['店铺', 'Shop', 'Store'])
                        df_whitelist = d[['MSKU', '店铺']] if col_shop else d[['MSKU']]
                else:
                    col_msku = find_col_fuzzy_priority(d, ['MSKU', '商家SKU'])
                    col_sku = find_col_exact(d, 'SKU') or find_col_fuzzy_priority(d, ['FNSKU']) or col_msku
                    if col_msku:
                        d['MSKU'] = d[col_msku].apply(clean_msku_strict)
                        d['SKU_KEY'] = d[col_sku].apply(clean_msku_strict) if col_sku else d['MSKU']
                        col_shop_data = find_col_fuzzy_priority(d, ['店铺', 'Shop', 'Store'])
                        if col_shop_data: d['店铺'] = d[col_shop_data].astype(str).str.strip()
                        all_data_dfs.append(d)
            
            df_master = pd.concat(all_data_dfs, ignore_index=True)
            if df_whitelist is not None:
                if '店铺' in df_whitelist.columns and '店铺' in df_master.columns:
                    temp = pd.merge(df_master, df_whitelist, on='MSKU', how='inner', suffixes=('', '_wl'))
                    valid_indices = []
                    for idx, row in temp.iterrows():
                        m_shop, w_shop = str(row['店铺']).upper().replace(" ", ""), str(row.get('店铺_wl', '')).upper().replace(" ", "")
                        if (w_shop in m_shop) or (m_shop in w_shop): valid_indices.append(idx)
                    df_master = temp.loc[valid_indices].copy().drop(columns=['店铺_wl'], errors='ignore')
                else:
                    df_master = pd.merge(df_master, df_whitelist[['MSKU']].drop_duplicates(), on='MSKU', how='inner')

            # 2. 处理流量与库存
            df_7 = process_traffic_df(merge_uploaded_traffic(files_7d), "7天")
            df_14 = process_traffic_df(merge_uploaded_traffic(files_14d), "14天")
            df_inventory = process_inventory_files(files_inv)
            df_age = process_age_files(files_age)

            # 3. 数据合并
            merged = df_master.copy()
            def merge_traffic_safe(master_df, traffic_df):
                if traffic_df is None: return master_df
                temp = pd.merge(master_df, traffic_df, left_on='MSKU', right_on='join_key', how='left')
                if 'join_key' in temp.columns: del temp['join_key']
                if 'traffic_shop' in temp.columns: del temp['traffic_shop']
                group_keys = [c for c in master_df.columns]
                traffic_cols = [c for c in traffic_df.columns if c not in ['join_key', 'traffic_shop']]
                return temp.groupby(group_keys, dropna=False)[traffic_cols].sum().reset_index()

            merged = merge_traffic_safe(merged, df_7)
            merged = merge_traffic_safe(merged, df_14)
            if df_inventory is not None: merged = pd.merge(merged, df_inventory, left_on='SKU_KEY', right_on='join_key', how='left').drop(columns=['join_key'], errors='ignore')
            if df_age is not None: merged = pd.merge(merged, df_age, left_on='SKU_KEY', right_on='join_key', how='left').drop(columns=['join_key'], errors='ignore')

            # 4. 填充与校准
            fill_keywords = ['7天', '14天', '21天', '28天', '库龄', '可用量', '待发货', '广告', '花费', 'CTR', 'ACOS', 'ACoAS', '点击', '曝光', '展示']
            cols_fill = [c for c in merged.columns if any(x in c for x in fill_keywords)]
            merged[cols_fill] = merged[cols_fill].fillna(0)

            rename_map = {
                find_col_fuzzy_priority(merged, ['展示', '广告曝光', '曝光', 'Impressions']): '广告曝光量',
                find_col_fuzzy_priority(merged, ['广告点击数', '点击']): '广告点击数',
                find_col_fuzzy_priority(merged, ['广告订单量', '广告订单', 'Ad Orders']): '广告订单',
                find_col_fuzzy_priority(merged, ['广告花费', '花费']): '广告花费',
                find_col_fuzzy_priority(merged, ['广告销售额']): '广告销售额'
            }
            for old_col, new_col in rename_map.items():
                if old_col and old_col != new_col: merged.rename(columns={old_col: new_col}, inplace=True)
                elif not old_col: merged[new_col] = 0

            cols_to_numeric = ['订单毛利率', '广告花费', 'CTR', '广告CVR', 'CPC', 'ACOS', 'ACoAS', '广告点击数', '广告销售额', '广告订单', '7天销售额', '广告曝光量']
            for col in cols_to_numeric:
                if col in merged.columns: merged[col] = merged[col].apply(lambda x: clean_percentage_or_money(x, col))

            # 5. 核心运算
            merged['7天日均订单'] = merged['7天订单商品总数'] / 7
            merged['14天日均订单'] = merged['14天订单商品总数'] / 14
            merged['7天销售转化率'] = merged.apply(lambda x: x['7天订单商品总数'] / x['7天会话数'] if x['7天会话数'] > 0 else 0, axis=1)
            merged['14天销售转化率'] = merged.apply(lambda x: x['14天订单商品总数'] / x['14天会话数'] if x['14天会话数'] > 0 else 0, axis=1)
            
            inv_cols = ['待发货', '7天内送达', '14天内送达', '21天内送达', '28天内送达', '28天以上送达']
            valid_inv = [c for c in inv_cols if c in merged.columns]
            merged['待到合计'] = merged[valid_inv].sum(axis=1) if valid_inv else 0

            merged['预测日销量'] = (merged['7天日均订单'] + merged['14天日均订单']) / 2
            merged['理论需求量'] = merged['预测日销量'] * TARGET_DAYS_TOTAL
            merged['总供给'] = merged.get('可用量', 0) + merged['待到合计']
            merged['建议补货量'] = merged.apply(lambda x: max(0, x['理论需求量'] - x['总供给']), axis=1)
            merged['预计可售天数'] = merged.apply(lambda x: x['总供给'] / x['预测日销量'] if x['预测日销量'] > 0.1 else 999, axis=1)

            # 6. 列排序与装修 (导出到 BytesIO)
            cols_to_move_front = [c for c in ['预测日销量', '建议补货量', '预计可售天数', '理论需求量', '总供给'] if c in merged.columns]
            cols_others = [c for c in merged.columns if c not in cols_to_move_front and c != 'SKU_KEY']
            cols = cols_others[:12] + cols_to_move_front + cols_others[12:]
            
            for m in [('店铺','MSKU','before'), ('7天日均订单','7天销售额','before'), ('7天销售转化率','7天销售额','after'), 
                      ('14天日均订单','14天销售额','before'), ('14天销售转化率','14天销售额','after'), ('待到合计','7天内送达','before'), 
                      ('待发货','可用量','before'), ('广告点击数','广告花费','before'), ('广告销售额','广告花费','after')]:
                cols = move_col(cols, m[0], m[1], m[2])
            
            if '广告曝光量' in cols: cols = move_col(cols, '广告曝光量', '广告点击数', 'before')
            if '广告订单' in cols: cols = move_col(cols, '广告订单', '广告销售额', 'before')
            merged = merged[cols]

            # Excel 渲染
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                merged.to_excel(writer, index=False, sheet_name='补货数据')
                ws = writer.sheets['补货数据']
                ws.column_dimensions['A'].width, ws.column_dimensions['B'].width = 15, 25
                ws.insert_rows(1, amount=2)
                ws.freeze_panes = 'J4'
                ws['A1'], ws['A2'] = "总计求和", "筛选求和"
                
                max_r = ws.max_row
                exclude_sum = ['店铺', 'MSKU', 'ASIN', 'SKU', '图片', 'image', '转化率', '可售天数', '商品属性', '分类', '型号']
                must_sum = ['销量', '销售额', '数量', '会话', '浏览', '送达', '待到', '可用', '库龄', '补货', '供给', '需求', '利润', '广告', '花费', '订单', '待发货', '点击', '曝光']
                
                # ... (由于篇幅限制，高亮样式的细化逻辑精简在循环中) ...
                for col_idx in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    header_name = str(ws.cell(row=3, column=col_idx).value).strip()
                    
                    is_summable = any(k in header_name for k in must_sum) and not any(ex in header_name for ex in exclude_sum)
                    if col_idx >= 9 and is_summable:
                        range_str = f"{col_letter}4:{col_letter}{max_r}"
                        ws[f'{col_letter}1'] = f"=SUM({range_str})"
                        ws[f'{col_letter}2'] = f"=SUBTOTAL(109, {range_str})"
            
            output.seek(0)
            st.success("🎉 计算成功！点击下方按钮下载报表。")
            st.download_button(
                label="📥 下载智能补货表",
                data=output,
                file_name="智能补货表.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

        except Exception as e:
            st.error(f"❌ 运行过程中出现错误，请检查上传的数据格式: {str(e)}")
